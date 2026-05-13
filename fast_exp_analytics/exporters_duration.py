from __future__ import annotations

from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    bad = ["\\", "/", "*", "?", ":", "[", "]"]
    s = str(name)
    for ch in bad:
        s = s.replace(ch, "_")
    return s[:max_len]


def _autosize_worksheet(ws, min_width: int = 10, max_width: int = 48):
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        max_len = 0
        for c in col_cells:
            try:
                value = "" if c.value is None else str(c.value)
            except Exception:
                value = ""
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            min_width, min(max_len + 2, max_width)
        )


def _style_header_row(ws, row_idx: int = 1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_df_to_sheet(ws, df: pd.DataFrame, title: str | None = None):
    row_start = 1

    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(size=13, bold=True, color="1F1F1F")
        row_start = 3

    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=row_start, column=j, value=col)

    _style_header_row(ws, row_start)

    for i, (_, row) in enumerate(df.iterrows(), start=row_start + 1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = None
            ws.cell(row=i, column=j, value=val)

    n_rows = row_start + len(df)
    n_cols = len(df.columns)

    ws.freeze_panes = ws.cell(row=row_start + 1, column=1)
    ws.auto_filter.ref = f"A{row_start}:{get_column_letter(n_cols)}{n_rows}"
    for row in ws.iter_rows(min_row=row_start + 1, max_row=n_rows):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    _apply_number_formats(ws, {c: i + 1 for i, c in enumerate(df.columns)}, n_rows)
    _autosize_worksheet(ws)


def _apply_number_formats(ws, header_map: dict[str, int], n_rows: int):
    percent_like_cols = [
        c
        for c in header_map
        if "MDE" in c or c.startswith("% ") or c.endswith(", %") or c == "Target MDE, %"
    ]
    integer_like_cols = [
        c
        for c in header_map
        if "Число дней" in c
        or "дней до target MDE" in c
        or c.endswith("_n_per_group")
        or c == "Число групп"
    ]
    float_like_cols = [
        "_units_per_day",
        "_eligible_per_day",
        "_smallest_group_share",
        "_alpha",
        "_power",
    ]

    for col in percent_like_cols:
        c = header_map[col]
        for r in range(2, n_rows + 1):
            ws.cell(r, c).number_format = "0.00"

    for col in integer_like_cols:
        c = header_map[col]
        for r in range(2, n_rows + 1):
            ws.cell(r, c).number_format = "#,##0"

    for col in float_like_cols:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = "#,##0.0000"


# =========================
# summary
def _find_primary_mde_columns(df: pd.DataFrame) -> list[str]:
    preferred = [
        "Списания MDE, %",
        "CPA MDE, %",
        "CTR MDE, %",
        "Цели MDE, %",
        "Amount per day MDE, %",
        "Goals per day MDE, %",
        "Показы MDE, %",
        "Клики MDE, %",
        "CPM MDE, %",
    ]
    existing = [c for c in preferred if c in df.columns]
    if existing:
        return existing[:5]

    fallback = [c for c in df.columns if "MDE" in c][:5]
    return fallback


def _find_primary_days_columns(df: pd.DataFrame) -> list[str]:
    preferred = [
        "Списания - дней до target MDE",
        "CPA — дней до target MDE",
        "CTR — дней до target MDE",
        "Цели — дней до target MDE",
        "Amount per day — дней до target MDE",
        "Goals per day — дней до target MDE",
        "Показы — дней до target MDE",
        "Клики — дней до target MDE",
        "CPM — дней до target MDE",
    ]
    existing = [c for c in preferred if c in df.columns]
    if existing:
        return existing[:5]

    fallback = [c for c in df.columns if "дней до target MDE" in c][:5]
    return fallback


def _fmt_num(x, decimals: int = 1) -> str:
    if pd.isna(x):
        return "н/д"
    return f"{float(x):.{decimals}f}".replace(".", ",")


def _fmt_int(x) -> str:
    if pd.isna(x):
        return "н/д"
    return f"{int(round(float(x))):,}".replace(",", " ")


def build_duration_manager_text(
    *,
    mde_by_days_df: Optional[pd.DataFrame] = None,
    days_for_target_mde_df: Optional[pd.DataFrame] = None,
    experiment_name: str | None = None,
    experiment_type: str | None = None,
    rollout_pct: float | None = None,
    recommended_days: int | None = None,
    comment: str | None = None,
) -> str:
    """
    Короткий manager-friendly текст.
    """
    lines: list[str] = []

    title = "Планирование длительности эксперимента"
    if experiment_name:
        title += f": {experiment_name}"
    lines.append(title)

    meta = []
    if experiment_type:
        meta.append(f"тип: {experiment_type.upper()}")
    if rollout_pct is not None:
        meta.append(f"выкатка: {_fmt_num(rollout_pct, 1)}%")
    if recommended_days is not None:
        meta.append(f"рекомендуемая длительность: {_fmt_int(recommended_days)} дн.")
    if meta:
        lines.append(" / ".join(meta))

    if comment:
        lines.append(comment)

    lines.append("")
    lines.append("Что видно по расчёту:")
    has_content = False

    if mde_by_days_df is not None and not mde_by_days_df.empty:
        has_content = True
        df = mde_by_days_df.copy()

        day_col = "Число дней экспа"
        if recommended_days is not None and day_col in df.columns:
            row = df.loc[df[day_col] == recommended_days]
            if row.empty:
                row = df.sort_values(day_col).tail(1)
        else:
            row = df.sort_values(day_col).tail(1) if day_col in df.columns else df.head(1)

        row = row.iloc[0]
        day_value = row[day_col] if day_col in row.index else recommended_days

        lines.append(
            f"- При длительности ~{_fmt_int(day_value)} дн. ожидаемый минимально детектируемый эффект составляет:"
        )

        for col in _find_primary_mde_columns(df):
            lines.append(f"  • {col.replace(' MDE, %', '')}: ~{_fmt_num(row[col], 1)}%")

    if days_for_target_mde_df is not None and not days_for_target_mde_df.empty:
        has_content = True
        df = days_for_target_mde_df.copy()

        target_col = "Target MDE, %"
        row = df.sort_values(target_col).head(1).iloc[0] if target_col in df.columns else df.head(1).iloc[0]

        target_val = row[target_col] if target_col in row.index else np.nan
        if not pd.isna(target_val):
            lines.append("")
            lines.append(f"- Чтобы дойти до чувствительности ~{_fmt_num(target_val, 1)}%, по ключевым метрикам нужно ориентироваться примерно на:")

            for col in _find_primary_days_columns(df):
                lines.append(f"  • {col.replace(' — дней до target MDE', '')}: ~{_fmt_int(row[col])} дн.")

    if not has_content:
        lines.append("- Недостаточно данных для формирования summary.")

    lines.append("")
    lines.append("Интерпретация:")
    lines.append("- Чем ниже MDE, тем более мелкий эффект мы можем надёжно увидеть в тесте")
    lines.append("- Если фактический бизнес-эффект ожидается меньше текущего MDE, эксперимент стоит катить дольше или расширять выкатку")
    lines.append("- Для ABC чувствительность обычно хуже, чем для AB, при той же выкладке и длительности, потому что трафик делится на большее число групп")

    return "\n".join(lines)


def _write_manager_summary_sheet(
    ws,
    summary_text: str,
    *,
    experiment_name: str | None = None,
):
    ws.sheet_view.showGridLines = False

    # Title
    ws["A1"] = "Сводка"
    ws["A1"].font = Font(size=16, bold=True, color="1F1F1F")

    if experiment_name:
        ws["A2"] = experiment_name
        ws["A2"].font = Font(size=11, italic=True, color="666666")

    # Text block
    start_row = 4
    paragraphs = summary_text.split("\n")

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="F4F8FB")
    text_fill = PatternFill("solid", fgColor="FFFFFF")
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )

    current_row = start_row
    for p in paragraphs:
        cell = ws.cell(current_row, 1, p)

        if current_row == start_row:
            cell.font = Font(size=12, bold=True, color="1F1F1F")
            cell.fill = title_fill
        elif p.endswith(":"):
            cell.font = Font(size=11, bold=True, color="1F1F1F")
            cell.fill = section_fill
        else:
            cell.font = Font(size=11, color="222222")
            cell.fill = text_fill

        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
        current_row += 1

    ws.column_dimensions["A"].width = 120
    ws.cell(current_row + 1, 1, "Примечание: расчёт основан на исторической вариативности метрик и ожидаемом размере групп. Это ориентир для планирования, а не гарантия результата")
    ws.cell(current_row + 1, 1).font = Font(size=10, italic=True, color="666666")
    ws.cell(current_row + 1, 1).alignment = Alignment(wrap_text=True, vertical="top")


def export_duration_results_to_excel(
    output_path: str | Path,
    *,
    mde_by_days_df: Optional[pd.DataFrame] = None,
    days_for_target_mde_df: Optional[pd.DataFrame] = None,
    experiment_name: str | None = None,
    experiment_type: str | None = None,
    rollout_pct: float | None = None,
    recommended_days: int | None = None,
    comment: str | None = None,
    summary_text: str | None = None,
) -> str:
    """
    Экспортирует:
      - manager_summary
      - mde_by_days
      - days_for_target_mde

    summary_text можно передать свой
    Если не передать, он будет сгенерирован автоматически
    """
    output_path = str(output_path)

    if mde_by_days_df is None and days_for_target_mde_df is None:
        raise ValueError("At least one of mde_by_days_df or days_for_target_mde_df must be provided.")

    if summary_text is None:
        summary_text = build_duration_manager_text(
            mde_by_days_df=mde_by_days_df,
            days_for_target_mde_df=days_for_target_mde_df,
            experiment_name=experiment_name,
            experiment_type=experiment_type,
            rollout_pct=rollout_pct,
            recommended_days=recommended_days,
            comment=comment,
        )

    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("manager_summary")
    _write_manager_summary_sheet(
        ws_summary,
        summary_text=summary_text,
        experiment_name=experiment_name,
    )

    if mde_by_days_df is not None:
        ws_mde = wb.create_sheet(_safe_sheet_name("mde_by_days"))
        title = "MDE по длительности эксперимента"
        if experiment_name:
            title += f" - {experiment_name}"
        _write_df_to_sheet(ws_mde, mde_by_days_df.copy(), title=title)

    if days_for_target_mde_df is not None:
        ws_days = wb.create_sheet(_safe_sheet_name("days_for_target_mde"))
        title = "Сколько дней нужно до target MDE"
        if experiment_name:
            title += f" - {experiment_name}"
        _write_df_to_sheet(ws_days, days_for_target_mde_df.copy(), title=title)

    wb.save(output_path)
    return output_path