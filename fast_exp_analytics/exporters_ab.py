from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    bad = ["\\", "/", "*", "?", ":", "[", "]"]
    s = str(name)
    for ch in bad:
        s = s.replace(ch, "_")
    return s[:max_len]


def _autosize_worksheet(ws, min_width: int = 10, max_width: int = 42):
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        max_len = 0
        for c in col_cells:
            try:
                value = "" if c.value is None else str(c.value)
            except Exception:
                value = ""
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            min_width, min(max_len + 2, max_width)
        )


def _style_header_row(ws, row_idx: int = 1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))

    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_result_cell_style(cell, value):
    v = str(value).lower() if value is not None else ""
    if v == "positive":
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(color="006100", bold=True)
    elif v == "negative":
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        cell.font = Font(color="9C0006", bold=True)
    elif v == "neutral":
        cell.fill = PatternFill("solid", fgColor="E7E6E6")
        cell.font = Font(color="666666", bold=False)


def _apply_number_formats(ws, header_map: dict[str, int], n_rows: int):
    percent_cols_fraction = ["rel_delta"]
    percent_cols_already_pct = ["avg_rel_delta", "power_now"]
    p_cols = ["p_value"]

    int_cols = [
        "number_samples",
        "number_samples_base",
        "number_samples_exp",
        "days_more_if_same_delta",
        "days_more_base",
        "days_more_exp",
        "n_required_per_group",
    ]

    decimal4_cols = [
        "value_base",
        "value_exp",
        "abs_delta",
        "mde",
        "avg_value_base",
        "avg_value_exp",
        "avg_abs_delta",
    ]

    for col in percent_cols_fraction:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = "0.0%"

    for col in percent_cols_already_pct:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = '0.0"%"'

    for col in p_cols:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = "0.0000"

    for col in int_cols:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = "#,##0"

    for col in decimal4_cols:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = "#,##0.0000"


def _write_df_to_sheet(ws, df: pd.DataFrame, title: str | None = None):
    row_start = 1

    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(size=13, bold=True, color="1F1F1F")
        row_start = 3

    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=row_start, column=j, value=col)

    _style_header_row(ws, row_start)

    for i, (_, row) in enumerate(df.iterrows(), start=row_start + 1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            if pd.isna(val):
                val = None
            ws.cell(row=i, column=j, value=val)

    n_rows = row_start + len(df)
    n_cols = len(df.columns)

    ws.freeze_panes = ws.cell(row=row_start + 1, column=1)
    ws.auto_filter.ref = f"A{row_start}:{get_column_letter(n_cols)}{n_rows}"

    header_map = {col: idx for idx, col in enumerate(df.columns, start=1)}
    _apply_number_formats(ws, header_map, n_rows)

    for row in ws.iter_rows(min_row=row_start + 1, max_row=n_rows):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    if "result" in header_map:
        c = header_map["result"]
        for r in range(row_start + 1, n_rows + 1):
            _apply_result_cell_style(ws.cell(r, c), ws.cell(r, c).value)

    if "p_value" in header_map:
        c = get_column_letter(header_map["p_value"])
        ws.conditional_formatting.add(
            f"{c}{row_start+1}:{c}{n_rows}",
            CellIsRule(
                operator="lessThan",
                formula=["0.05"],
                fill=PatternFill("solid", fgColor="FFF2CC"),
            ),
        )

    if "rel_delta" in header_map:
        c = get_column_letter(header_map["rel_delta"])
        ws.conditional_formatting.add(
            f"{c}{row_start+1}:{c}{n_rows}",
            CellIsRule(
                operator="greaterThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor="E2F0D9"),
            ),
        )
        ws.conditional_formatting.add(
            f"{c}{row_start+1}:{c}{n_rows}",
            CellIsRule(
                operator="lessThan",
                formula=["0"],
                fill=PatternFill("solid", fgColor="FCE4D6"),
            ),
        )

    _autosize_worksheet(ws)


def export_ab_results_to_excel(
    df_result_ab: pd.DataFrame,
    output_path: str | Path,
    *,
    experiment_desc: str | None = None,
    exp_id: str | int | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
) -> str:
    output_path = str(output_path)

    df = df_result_ab.copy()

    preferred_cols = [
        "metric_name",
        "metric_type",
        "value_base",
        "value_exp",
        "abs_delta",
        "rel_delta",
        "p_value",
        "result",
        "mde",
        "number_samples",
        "number_samples_base",
        "number_samples_exp",
        "avg_value_base",
        "avg_value_exp",
        "avg_abs_delta",
        "avg_rel_delta",
        "days_more_if_same_delta",
        "days_more_base",
        "days_more_exp",
        "n_required_per_group",
        "power_now",
        "direction",
    ]
    existing_cols = [c for c in preferred_cols if c in df.columns] + [
        c for c in df.columns if c not in preferred_cols
    ]
    df = df[existing_cols].copy()

    wb = Workbook()
    wb.remove(wb.active)

    ws_all = wb.create_sheet("ab_metrics")
    title = "AB metrics"
    if experiment_desc:
        title += f" - {experiment_desc}"
    _write_df_to_sheet(ws_all, df, title=title)

    ws_meta = wb.create_sheet("summary")
    meta_df = pd.DataFrame(
        {
            "field": ["experiment_desc", "exp_id", "date_from", "date_to", "rows_exported"],
            "value": [experiment_desc, exp_id, date_from, date_to, len(df)],
        }
    )
    _write_df_to_sheet(ws_meta, meta_df, title="Summary")

    # summary первым листом
    wb._sheets.remove(ws_meta)
    wb._sheets.insert(0, ws_meta)
    wb.active = 0

    wb.save(output_path)
    return output_path