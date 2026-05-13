from __future__ import annotations

from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

def _safe_sheet_name(name: str, max_len: int = 31) -> str:
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        name = str(name).replace(ch, "_")
    return str(name)[:max_len]

def _autosize_worksheet(ws, min_width: int = 10, max_width: int = 42):
    for col_cells in ws.columns:
        col_idx = col_cells[0].column
        max_len = max(len("" if c.value is None else str(c.value)) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(min_width, min(max_len + 2, max_width))

def _style_header_row(ws, row_idx: int = 1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def _apply_result_cell_style(cell, value):
    v = str(value).lower() if value is not None else ""
    if v == "positive":
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(color="006100", bold=True)
    elif v == "negative":
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        cell.font = Font(color="9C0006", bold=True)
    elif v == "neutral":
        cell.fill = PatternFill("solid", fgColor="E7E6E6")
        cell.font = Font(color="666666")

def _apply_number_formats(ws, header_map: dict[str, int], n_rows: int):
    percent_fraction = ["rel_delta"]
    percent_plain = ["avg_rel_delta", "power_now"]
    p_cols = ["p_value", "p_value_adj"]
    int_cols = ["number_samples", "number_samples_base", "number_samples_exp", "days_more_if_same_delta", "days_more_base", "days_more_exp", "n_required_per_group"]
    decimal4_cols = ["value_base", "value_exp", "abs_delta", "mde", "avg_value_base", "avg_value_exp", "avg_abs_delta"]
    for col in percent_fraction:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = "0.0%"
    for col in percent_plain:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = '0.0"%"'
    for col in p_cols:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = "0.0000"
    for col in int_cols:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = "#,##0"
    for col in decimal4_cols:
        if col in header_map:
            c = header_map[col]
            for r in range(2, n_rows + 1):
                ws.cell(r, c).number_format = "#,##0.0000"

def _write_df_to_sheet(ws, df: pd.DataFrame, title: str | None = None):
    row_start = 1
    if title:
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = Font(size=13, bold=True, color="1F1F1F")
        row_start = 3

    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=row_start, column=j, value=col)
    _style_header_row(ws, row_start)

    for i, (_, row) in enumerate(df.iterrows(), start=row_start + 1):
        for j, col in enumerate(df.columns, start=1):
            val = None if pd.isna(row[col]) else row[col]
            ws.cell(row=i, column=j, value=val)

    n_rows = row_start + len(df)
    n_cols = len(df.columns)
    ws.freeze_panes = ws.cell(row=row_start + 1, column=1)
    ws.auto_filter.ref = f"A{row_start}:{get_column_letter(n_cols)}{n_rows}"
    header_map = {col: idx for idx, col in enumerate(df.columns, start=1)}
    _apply_number_formats(ws, header_map, n_rows)

    for row in ws.iter_rows(min_row=row_start + 1, max_row=n_rows):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    for result_col in ["result", "result_adj"]:
        if result_col in header_map:
            c = header_map[result_col]
            for r in range(row_start + 1, n_rows + 1):
                _apply_result_cell_style(ws.cell(r, c), ws.cell(r, c).value)

    for pcol in ["p_value", "p_value_adj"]:
        if pcol in header_map:
            c = get_column_letter(header_map[pcol])
            ws.conditional_formatting.add(
                f"{c}{row_start+1}:{c}{n_rows}",
                CellIsRule(operator="lessThan", formula=["0.05"], fill=PatternFill("solid", fgColor="FFF2CC")),
            )

    if "rel_delta" in header_map:
        c = get_column_letter(header_map["rel_delta"])
        ws.conditional_formatting.add(f"{c}{row_start+1}:{c}{n_rows}", CellIsRule(operator="greaterThan", formula=["0"], fill=PatternFill("solid", fgColor="E2F0D9")))
        ws.conditional_formatting.add(f"{c}{row_start+1}:{c}{n_rows}", CellIsRule(operator="lessThan", formula=["0"], fill=PatternFill("solid", fgColor="FCE4D6")))

    _autosize_worksheet(ws)

def _build_summary_df(df_result_abc: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for pair, sub in df_result_abc.groupby("pair", dropna=False):
        result_col = "result_adj" if "result_adj" in sub.columns else "result"
        cnt_pos = int((sub[result_col] == "positive").sum()) if result_col in sub.columns else 0
        cnt_neg = int((sub[result_col] == "negative").sum()) if result_col in sub.columns else 0
        cnt_neu = int((sub[result_col] == "neutral").sum()) if result_col in sub.columns else 0
        best = sub.assign(abs_rel_delta=sub["rel_delta"].abs()).sort_values(["p_value_adj" if "p_value_adj" in sub.columns else "p_value", "abs_rel_delta"], ascending=[True, False]).head(3)
        top_metrics = ", ".join(f"{r.metric_name} ({r.rel_delta:.1%})" for r in best.itertuples() if pd.notna(r.rel_delta))
        rows.append({"pair": pair, "n_metrics": len(sub), "positive_cnt": cnt_pos, "negative_cnt": cnt_neg, "neutral_cnt": cnt_neu, "top_metrics": top_metrics})
    return pd.DataFrame(rows)

def export_abc_results_to_excel(df_result_abc: pd.DataFrame, output_path: str | Path, *, experiment_desc: str | None = None, exp_id: str | int | None = None, date_from: str | None = None, date_to: str | None = None) -> str:
    output_path = str(output_path)
    df = df_result_abc.copy()
    wb = Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("summary")
    summary_df = _build_summary_df(df)
    summary_title = "Метрики теста" + (f" - {experiment_desc}" if experiment_desc else "")
    _write_df_to_sheet(ws_summary, summary_df, title=summary_title)

    meta_row = len(summary_df) + 6
    meta = [("experiment_desc", experiment_desc), ("exp_id", exp_id), ("date_from", date_from), ("date_to", date_to), ("rows_exported", len(df))]
    for i, (k, v) in enumerate(meta, start=meta_row):
        ws_summary.cell(i, 1, k)
        ws_summary.cell(i, 2, "" if v is None else str(v))
        ws_summary.cell(i, 1).font = Font(bold=True)
    _autosize_worksheet(ws_summary)

    ws_all = wb.create_sheet("all_metrics")
    _write_df_to_sheet(ws_all, df, title=("All ABC metrics" + (f" - {experiment_desc}" if experiment_desc else "")))

    if "pair" in df.columns:
        for pair in ["A_vs_B", "A_vs_C", "B_vs_C"]:
            sub = df[df["pair"] == pair].copy()
            if sub.empty:
                continue
            ws = wb.create_sheet(_safe_sheet_name(pair))
            _write_df_to_sheet(ws, sub, title=f"Pair: {pair}")

    wb.save(output_path)
    return output_path
